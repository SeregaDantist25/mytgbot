# -*- coding: utf-8 -*-

from io import BytesIO

from openpyxl import load_workbook

from services.defect_act_service import DefectActError, generate_defect_act
from services.defect_profiles import build_defect_rows, detect_defect_profile


def test_generate_defect_act_fills_template():
    result = generate_defect_act(
        {
            "act_number": "4.10",
            "act_date": "27.08.2026",
            "ship": "Славянская",
            "order_number": "24-01",
            "repair_item": "4.10",
            "manager": "С. В. Бачурин",
            "equipment": "Воздушный клапан Ду-25",
            "repair_category": "Текущий ремонт",
            "work_summary": "Дефектация двух воздушных клапанов",
            "basis": "Ремонтная ведомость",
            "rows": [{
                "defect": "Износ уплотнительных поверхностей",
                "work": "Разобрать, дефектовать, восстановить и испытать",
                "unit": "шт.",
                "qty": 2,
            }],
            "conclusion": "Клапаны подлежат ремонту с последующим испытанием.",
        }
    )

    sheet = load_workbook(BytesIO(result))["Акт дефектации"]
    assert sheet["A6"].value == "№ 4.10"
    assert sheet["A8"].value == "Объект ремонта: Славянская"
    assert sheet["A14"].value == "1"
    assert sheet["E14"].value == "2"
    assert sheet["A15"].value == ""
    assert "{{" not in " ".join(
        str(cell.value) for row in sheet.iter_rows() for cell in row if cell.value
    )


def test_generate_defect_act_expands_table():
    result = generate_defect_act({"rows": [{"defect": str(i)} for i in range(15)]})
    sheet = load_workbook(BytesIO(result))["Акт дефектации"]
    assert sheet["B28"].value == "14"


def test_pipeline_profile_uses_flat_rows():
    profile = detect_defect_profile("Трубопровод охлаждения Ø57 мм")
    rows = build_defect_rows("Трубопровод", ["Сквозная коррозия"], "Замена", profile)
    assert profile == "pipeline"
    assert rows[0]["num"] == "1"
    assert "остаточную толщину" in rows[0]["work"]


def test_quantity_is_taken_from_repair_item():
    rows = build_defect_rows(
        "Трубопровод", ["Коррозия"], "Замена", "pipeline", quantity="2 шт."
    )
    assert rows[0]["qty"] == "2"
    assert rows[0]["unit"] == "шт."


def test_gear_profile_uses_hierarchical_rows():
    profile = detect_defect_profile("Редуктор брашпиля")
    rows = build_defect_rows("Редуктор брашпиля", ["Износ зубьев"], "Разобрать", profile)
    assert profile == "deck_machinery"
    assert [row["num"] for row in rows] == ["1", "1.1"]
