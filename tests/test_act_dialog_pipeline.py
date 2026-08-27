# -*- coding: utf-8 -*-

from io import BytesIO

from openpyxl import load_workbook

from ai.act_dialog import build_act_file


def test_dialog_session_builds_complete_xlsx():
    file_bytes, work_volume = build_act_file({
        "item_number": "6.1",
        "ship": "Славянская",
        "equipment": "Редуктор брашпиля Б-5",
        "equipment_type": "mechanism",
        "pump_type": None,
        "gosts": [],
        "defects": ["Износ зубьев", "Повышенный люфт вала"],
        "repair_type": "Средний ремонт",
        "extra_info": "Количество — 1 комплект",
        "order_number": "24-01",
        "manager_name": "С. В. Бачурин",
        "contractor_name": "С. В. Бачурин",
        "item_quantity": "1 компл.",
    })

    assert file_bytes.startswith(b"PK")
    assert work_volume
    sheet = load_workbook(BytesIO(file_bytes))["Акт дефектации"]
    values = " ".join(
        str(cell.value) for row in sheet.iter_rows() for cell in row if cell.value
    )
    assert "№ 6.1" in values
    assert "Славянская" in values
    assert "№ заказа: 24-01" in values
    assert "Редуктор брашпиля Б-5" in values
    assert "1.1" in values and "1.2" in values
    assert "{{" not in values
