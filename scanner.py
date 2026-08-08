# -*- coding: utf-8 -*-
"""
Сканер папки repair_docs.

- Парсит excel-файлы ремонтных ведомостей в таблицу repair_items.
- Обрабатывает PDF-договоры (проверка валидности, отправка на утверждение).
- Определяет судно по имени файла.

Соглашение об именах файлов:
  Ремведомость_<Судно>.xlsx   — ремонтная ведомость
  Договор_<Судно>.pdf         — скан-копия договора
  <Судно>_<что-то>.pdf        — договор (судно определяется по имени)
"""

import os
import re
import shutil
import sqlite3
import time
from datetime import datetime

import openpyxl

import db

REPAIR_DOCS_DIR = "repair_docs"
DOCUMENTS_DIR = os.path.join("data", "documents")

# Расширения
XLSX_EXT = (".xlsx", ".xls")
PDF_EXT = ".pdf"

# Папка для обработанных файлов (чтобы не сканировать повторно)
PROCESSED_DIR = os.path.join(REPAIR_DOCS_DIR, "_processed")


def _normalize(name):
    """Приводит имя к нижнему регистру без лишних пробелов."""
    return re.sub(r"\s+", " ", name.strip().lower())


def _norm_item_number(num):
    """Нормализует номер пункта: убирает завершающую точку и пробелы."""
    return num.strip().rstrip(".")


def detect_ship_from_filename(filename):
    """Определяет судно по имени файла. Возвращает имя судна или None."""
    base = os.path.splitext(os.path.basename(filename))[0]
    base = _normalize(base)
    # Убираем служебные префиксы
    base = re.sub(r"^(ремведомость|ведомость|договор|договор_|скан)\s*[_\-\s]*", "", base)
    # Ищем судно в базе по вхождению в имя файла
    for ship in db.get_ships():
        ship_norm = _normalize(ship["name"])
        if ship_norm and ship_norm in base:
            return ship["name"]
    return None


def _is_repair_list(filename):
    return filename.lower().endswith(XLSX_EXT)


def _is_contract(filename):
    return filename.lower().endswith(PDF_EXT)


def _is_processed(filename):
    """Проверяет, что файл уже обработан (лежит в _processed)."""
    return PROCESSED_DIR in os.path.normpath(filename)


def _move_to_processed(src):
    """Перемещает обработанный файл в папку _processed."""
    if not os.path.exists(PROCESSED_DIR):
        os.makedirs(PROCESSED_DIR)
    dst = os.path.join(PROCESSED_DIR, os.path.basename(src))
    # Если файл с таким именем уже есть — добавляем суффикс
    if os.path.exists(dst):
        base, ext = os.path.splitext(dst)
        dst = f"{base}_{int(time.time())}{ext}"
    shutil.move(src, dst)
    return dst


def parse_repair_list(path):
    """
    Парсит excel ремонтной ведомости в список пунктов.
    Возвращает список dict {item_number, description, quantity}.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[0]

    items = []
    current_section = None

    for row in ws.iter_rows():
        num_cell = row[0].value if len(row) > 0 else None
        name_cell = row[1].value if len(row) > 1 else None
        qty_cell = row[3].value if len(row) > 3 else None  # колонка D — объём/кол-во

        # Пропускаем пустые строки
        if num_cell is None and name_cell is None:
            continue

        # Заголовок раздела: "Раздел X. Название"
        if name_cell and isinstance(name_cell, str) and "Раздел" in name_cell:
            current_section = name_cell.strip()
            continue

        # Пропускаем строку заголовка таблицы
        if name_cell and isinstance(name_cell, str) and "Наименование работ" in name_cell:
            continue

        # Пункт: есть номер в колонке A и описание в колонке B
        if num_cell is not None and name_cell and isinstance(name_cell, str):
            num = str(num_cell).strip()
            desc = name_cell.strip()
            # Пропускаем служебные строки (нумерация колонок "1", "2")
            if num in ("1", "2") and len(desc) <= 2:
                continue
            if num and desc:
                items.append({
                    "item_number": num,
                    "description": desc,
                    "quantity": str(qty_cell).strip() if qty_cell is not None else None,
                    "section": current_section,
                })

    return items


def parse_exclusions(path):
    """
    Парсит смету-ведомость и извлекает исключения.
    Возвращает dict:
      {
        "partial": {item_number: quantity},   # исключено частично (кол-во к вычету)
        "full": [item_number, ...],           # исключено полностью
        "extra": [ {item_number, description}, ... ]  # дополнительные работы
      }
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    # Ищем лист с исключениями (Оригинал или Переработанная)
    ws = None
    for name in ("Оригинал", "Переработанная"):
        if name in wb.sheetnames:
            ws = wb[name]
            break
    if ws is None:
        ws = wb.worksheets[0]

    result = {"partial": {}, "full": [], "extra": []}
    mode = None  # None | "partial" | "full" | "extra"

    for row in ws.iter_rows():
        num_cell = row[0].value if len(row) > 0 else None
        name_cell = row[1].value if len(row) > 1 else None
        qty_cell = row[3].value if len(row) > 3 else None

        # Заголовки разделов могут быть в колонке A или B
        header = None
        if name_cell and isinstance(name_cell, str):
            header = name_cell.strip()
        elif num_cell and isinstance(num_cell, str):
            header = num_cell.strip()

        if header:
            t = header
            if "ИСКЛЮЧЕННЫЕ РАБОТЫ" in t.upper():
                continue
            if "Исключенные частично" in t:
                mode = "partial"
                continue
            if "Исключенные полностью" in t:
                mode = "full"
                continue
            if "Дополнительные работы" in t or "Дополнительный заказ" in t:
                mode = "extra"
                continue
            if t in ("Итого", "НДС 20 %") or t.startswith("Всего"):
                continue

        if num_cell is None:
            continue
        num = str(num_cell).strip()
        if not num:
            continue

        if mode == "partial":
            if name_cell and isinstance(name_cell, str):
                result["partial"][_norm_item_number(num)] = qty_cell
        elif mode == "full":
            if name_cell and isinstance(name_cell, str):
                result["full"].append(_norm_item_number(num))
        elif mode == "extra":
            if name_cell and isinstance(name_cell, str):
                result["extra"].append({
                    "item_number": f"Доп-{num}",
                    "description": name_cell.strip(),
                })

    return result


def apply_exclusions(items, exclusions):
    """
    Применяет исключения к списку пунктов ремонтной ведомости.
    items — список dict {item_number, description, quantity}.
    Возвращает новый список с учётом исключений.
    """
    partial = exclusions.get("partial", {})
    full = set(exclusions.get("full", []))
    extra = exclusions.get("extra", [])

    result = []
    for it in items:
        num = _norm_item_number(it["item_number"])
        # Копируем, чтобы не мутировать исходный список
        new_item = dict(it)
        # Исключено полностью — пропускаем
        if num in full:
            continue
        # Исключено частично — вычитаем количество
        if num in partial:
            excl_qty = partial[num]
            cur_qty = new_item.get("quantity")
            try:
                cur = float(str(cur_qty).replace(",", ".")) if cur_qty else 0
                excl = float(str(excl_qty).replace(",", ".")) if excl_qty else 0
                new_qty = cur - excl
                new_item["quantity"] = str(new_qty) if new_qty > 0 else "0"
                new_item["status"] = "reduced"
            except (ValueError, TypeError):
                new_item["status"] = "reduced"
        result.append(new_item)

    # Добавляем дополнительные работы
    for e in extra:
        result.append({
            "item_number": e["item_number"],
            "description": e["description"],
            "quantity": None,
            "status": "extra",
        })

    return result


def process_repair_list(path, ship_name):
    """Обрабатывает ремонтную ведомость: парсит и сохраняет пункты."""
    ship = db.get_ship_by_name(ship_name)
    if not ship:
        return False, f"Судно «{ship_name}» не найдено в базе."
    items = parse_repair_list(path)
    if not items:
        return False, f"В файле {os.path.basename(path)} не найдено пунктов."
    db.replace_repair_items(ship["ship_id"], items)
    return True, f"Ремонтная ведомость: загружено {len(items)} пунктов."


def _is_smeta(path):
    """Определяет, является ли excel-файл сметой с исключениями."""
    try:
        ex = parse_exclusions(path)
        return bool(ex["partial"] or ex["full"] or ex["extra"])
    except Exception:
        return False


def process_exclusions(path, ship_name):
    """Применяет исключения из сметы к уже загруженной ремонтной ведомости судна."""
    ship = db.get_ship_by_name(ship_name)
    if not ship:
        return False, f"Судно «{ship_name}» не найдено в базе."

    ex = parse_exclusions(path)
    if not (ex["partial"] or ex["full"] or ex["extra"]):
        return False, f"В файле {os.path.basename(path)} не найдено исключений."

    current = db.get_repair_items(ship["ship_id"])
    if not current:
        return False, "Сначала загрузите ремонтную ведомость (нет пунктов для применения исключений)."

    items = [{
        "item_number": it["item_number"],
        "description": it["description"],
        "quantity": it.get("quantity"),
        "status": it.get("status", "active"),
    } for it in current]

    result = apply_exclusions(items, ex)
    db.replace_repair_items(ship["ship_id"], result)

    n_partial = len(ex["partial"])
    n_full = len(ex["full"])
    n_extra = len(ex["extra"])
    return True, (f"Исключения применены: частично {n_partial}, полностью {n_full}, "
                  f"доп. работ {n_extra}. Итог: {len(result)} пунктов.")


def validate_pdf(path):
    """Проверяет, что PDF валидный и непустой."""
    try:
        with open(path, "rb") as f:
            header = f.read(1024)
        if not header.startswith(b"%PDF"):
            return False, "Файл не является PDF (нет сигнатуры %PDF)."
        size = os.path.getsize(path)
        if size < 100:
            return False, "PDF-файл пустой или слишком мал."
        return True, "PDF валидный."
    except Exception as e:
        return False, f"Ошибка чтения PDF: {e}"


def process_contract(path, ship_name, uploaded_by=None):
    """Обрабатывает договор: проверяет валидность, сохраняет, отправляет на утверждение."""
    ship = db.get_ship_by_name(ship_name)
    if not ship:
        return False, f"Судно «{ship_name}» не найдено в базе."

    ok, msg = validate_pdf(path)
    if not ok:
        return False, f"Договор отклонён: {msg}"

    # Сохраняем файл в data/documents/<судно>/договор/
    ship_dir = os.path.join(DOCUMENTS_DIR, ship_name, "договор")
    if not os.path.exists(ship_dir):
        os.makedirs(ship_dir)
    dst = os.path.join(ship_dir, os.path.basename(path))
    if os.path.exists(dst):
        base, ext = os.path.splitext(dst)
        dst = f"{base}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{ext}"
    shutil.copy(path, dst)

    # Сохраняем в БД (договор не утверждён — ждёт одобрения)
    doc_id, err = db.add_document(
        ship["ship_id"], db.DOC_CONTRACT, dst, uploaded_by, approved=0
    )
    if err:
        return False, err
    return True, f"Договор сохранён и отправлен на утверждение (id={doc_id})."


def scan_repair_docs(uploaded_by=None):
    """
    Сканирует папку repair_docs, обрабатывает новые файлы.
    Возвращает список сообщений о результатах.
    """
    if not os.path.exists(REPAIR_DOCS_DIR):
        return ["Папка repair_docs не найдена."]

    messages = []
    for filename in sorted(os.listdir(REPAIR_DOCS_DIR)):
        path = os.path.join(REPAIR_DOCS_DIR, filename)
        if not os.path.isfile(path):
            continue
        if _is_processed(path):
            continue

        ship_name = detect_ship_from_filename(filename)
        if not ship_name:
            messages.append(f"⚠️ Не удалось определить судно по имени файла: {filename}")
            continue

        try:
            if _is_repair_list(filename):
                if _is_smeta(path):
                    ok, msg = process_exclusions(path, ship_name)
                else:
                    ok, msg = process_repair_list(path, ship_name)
                messages.append(f"{'✅' if ok else '❌'} {msg}")
                if ok:
                    _move_to_processed(path)
            elif _is_contract(filename):
                ok, msg = process_contract(path, ship_name, uploaded_by)
                messages.append(f"{'✅' if ok else '❌'} {msg}")
                if ok:
                    _move_to_processed(path)
            else:
                messages.append(f"⚠️ Неподдерживаемый формат файла: {filename}")
        except Exception as e:
            messages.append(f"❌ Ошибка обработки {filename}: {e}")

    if not messages:
        messages.append("Новых файлов для обработки не найдено.")
    return messages
