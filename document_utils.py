# -*- coding: utf-8 -*-
"""
Утилиты для работы с документами.
Включает функции для замены, конвертации в PDF и других операций.
"""

import os
from models import SessionLocal, Document


def handle_document_replace(document_id, new_file_path, user_id):
    """
    Заменить draft-документ новым файлом.
    
    Args:
        document_id: ID документа для замены
        new_file_path: Путь к новому файлу
        user_id: ID пользователя, выполняющего замену
    
    Returns:
        bool: True если успешно, False если ошибка
    """
    session = SessionLocal()
    try:
        doc = session.query(Document).filter_by(id=document_id).first()
        
        if not doc:
            return False
        
        if doc.status != "draft":
            return False
        
        # Обновляем путь к файлу
        doc.file_ref = new_file_path
        doc.file_type = os.path.splitext(new_file_path)[1].lower()
        
        session.commit()
        return True
    finally:
        session.close()


def convert_to_pdf(file_path, file_type):
    """
    Конвертировать документ в PDF.
    
    Args:
        file_path: Путь к исходному файлу
        file_type: Тип файла (.docx, .xlsx, .pdf)
    
    Returns:
        str: Путь к PDF-файлу или None если ошибка
    """
    
    if file_type == ".pdf":
        # Уже PDF, возвращаем как есть
        return file_path
    
    if file_type == ".docx":
        return _convert_docx_to_pdf(file_path)
    elif file_type == ".xlsx":
        return _convert_xlsx_to_pdf(file_path)
    
    return None


def _convert_docx_to_pdf(docx_path):
    """Конвертировать DOCX в PDF."""
    try:
        from docx import Document as DocxDocument
        from reportlab.lib.pagesizes import letter
        from reportlab.pdfgen import canvas
        from reportlab.lib.utils import ImageReader
        
        # Читаем DOCX
        doc = DocxDocument(docx_path)
        
        # Создаём PDF
        pdf_path = docx_path.replace(".docx", ".pdf")
        c = canvas.Canvas(pdf_path, pagesize=letter)
        
        y = 750
        for para in doc.paragraphs:
            if para.text.strip():
                c.drawString(50, y, para.text[:100])  # Ограничиваем длину строки
                y -= 20
                if y < 50:
                    c.showPage()
                    y = 750
        
        c.save()
        return pdf_path
    except Exception as e:
        print(f"Ошибка при конвертации DOCX: {e}")
        return None


def _convert_xlsx_to_pdf(xlsx_path):
    """Конвертировать XLSX в PDF."""
    try:
        from openpyxl import load_workbook
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.pdfgen import canvas
        
        # Читаем XLSX
        wb = load_workbook(xlsx_path)
        ws = wb.active
        
        # Создаём PDF
        pdf_path = xlsx_path.replace(".xlsx", ".pdf")
        c = canvas.Canvas(pdf_path, pagesize=landscape(letter))
        
        y = 750
        for row in ws.iter_rows(values_only=True):
            row_text = " | ".join(str(cell) if cell else "" for cell in row)
            if row_text.strip():
                c.drawString(50, y, row_text[:150])  # Ограничиваем длину строки
                y -= 15
                if y < 50:
                    c.showPage()
                    y = 750
        
        c.save()
        return pdf_path
    except Exception as e:
        print(f"Ошибка при конвертации XLSX: {e}")
        return None


def handle_document_approve_with_pdf(document_id, user_id):
    """
    Утвердить документ и конвертировать в PDF.
    
    Args:
        document_id: ID документа
        user_id: ID пользователя, утверждающего документ
    
    Returns:
        bool: True если успешно, False если ошибка
    """
    session = SessionLocal()
    try:
        doc = session.query(Document).filter_by(id=document_id).first()
        
        if not doc or doc.status != "draft":
            return False
        
        # Конвертируем в PDF
        pdf_path = convert_to_pdf(doc.file_ref, doc.file_type)
        
        if pdf_path:
            doc.file_ref = pdf_path
            doc.file_type = ".pdf"
        
        # Утверждаем документ
        doc.status = "approved"
        
        session.commit()
        return True
    finally:
        session.close()
