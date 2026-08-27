# -*- coding: utf-8 -*-
"""Генерация актов дефектации по утверждённому XLSX-шаблону."""

from __future__ import annotations

from copy import copy
from io import BytesIO
from pathlib import Path
import re
from typing import Any, Mapping, Sequence

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils.cell import get_column_letter, range_boundaries


DEFAULT_TEMPLATE = Path("templates/defect_act_template.xlsx")
TEMPLATE_DEFECT_ROWS = 12
PLACEHOLDER_RE = re.compile(r"{{\s*([a-zA-Z0-9_]+)\s*}}")


class DefectActError(ValueError):
    """Ошибка входных данных или шаблона акта."""


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _build_context(data: Mapping[str, Any]) -> dict[str, str]:
    context = {
        "act_number": _text(data.get("act_number")),
        "act_date": _text(data.get("act_date")),
        "ship": _text(data.get("ship")),
        "order_number": _text(data.get("order_number")),
        "repair_item": _text(data.get("repair_item")),
        "manager": _text(data.get("manager")),
        "equipment": _text(data.get("equipment")),
        "repair_category": _text(data.get("repair_category")),
        "work_summary": _text(data.get("work_summary")),
        "basis": _text(data.get("basis")),
        "conclusion": _text(data.get("conclusion")),
        "contractor_name": _text(data.get("contractor_name")),
        "qc_name": _text(data.get("qc_name")),
        "customer_name": _text(data.get("customer_name")),
        "ship_representative": _text(data.get("ship_representative")),
        "sign_date": _text(data.get("sign_date") or data.get("act_date")),
    }

    rows = data.get("rows") or []
    if not isinstance(rows, Sequence) or isinstance(rows, (str, bytes)):
        raise DefectActError("Поле rows должно быть списком позиций")
    for index in range(1, max(TEMPLATE_DEFECT_ROWS, len(rows)) + 1):
        row = rows[index - 1] if index <= len(rows) else {}
        if not isinstance(row, Mapping):
            raise DefectActError(f"Позиция {index} должна быть объектом")
        context[f"row_{index}_num"] = _text(row.get("num") or (index if row else ""))
        context[f"row_{index}_defect"] = _text(row.get("defect"))
        context[f"row_{index}_work"] = _text(row.get("work"))
        context[f"row_{index}_unit"] = _text(row.get("unit"))
        context[f"row_{index}_qty"] = _text(row.get("qty"))
        context[f"row_{index}_note"] = _text(row.get("note"))
    return context


def _replace_placeholders(value: Any, context: Mapping[str, str]) -> Any:
    if not isinstance(value, str) or "{{" not in value:
        return value
    return PLACEHOLDER_RE.sub(lambda match: context.get(match.group(1), ""), value)


def generate_defect_act(
    data: Mapping[str, Any], template_path: str | Path = DEFAULT_TEMPLATE
) -> bytes:
    """Заполняет XLSX-шаблон и возвращает готовый файл в байтах."""
    template = Path(template_path)
    if not template.is_file():
        raise DefectActError(f"Шаблон акта не найден: {template}")

    workbook = load_workbook(template)
    if "Акт дефектации" not in workbook.sheetnames:
        raise DefectActError("В шаблоне отсутствует лист «Акт дефектации»")

    context = _build_context(data)
    sheet = workbook["Акт дефектации"]
    rows = data.get("rows") or []
    if len(rows) > TEMPLATE_DEFECT_ROWS:
        extra_count = len(rows) - TEMPLATE_DEFECT_ROWS
        lower_merges = [
            str(cell_range)
            for cell_range in sheet.merged_cells.ranges
            if cell_range.min_row >= 27
        ]
        for cell_range in lower_merges:
            sheet.unmerge_cells(cell_range)
        sheet.insert_rows(26, amount=extra_count)
        for cell_range in lower_merges:
            min_col, min_row, max_col, max_row = range_boundaries(cell_range)
            shifted = (
                f"{get_column_letter(min_col)}{min_row + extra_count}:"
                f"{get_column_letter(max_col)}{max_row + extra_count}"
            )
            sheet.merge_cells(shifted)
        for row_index in range(26, 26 + extra_count):
            source_row = 25
            sheet.row_dimensions[row_index].height = sheet.row_dimensions[source_row].height
            for column in range(1, 7):
                source = sheet.cell(source_row, column)
                target = sheet.cell(row_index, column)
                if source.has_style:
                    target._style = copy(source._style)
                field = ("num", "defect", "work", "unit", "qty", "note")[column - 1]
                row_number = row_index - 13
                target.value = "{{row_%d_%s}}" % (row_number, field)
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            cell.value = _replace_placeholders(cell.value, context)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def save_defect_act(
    output_path: str | Path,
    data: Mapping[str, Any],
    template_path: str | Path = DEFAULT_TEMPLATE,
) -> Path:
    """Создаёт акт и сохраняет его по указанному пути."""
    target = Path(output_path)
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_bytes(generate_defect_act(data, template_path))
    return target
